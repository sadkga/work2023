# coding=utf-8
"""
 -- @ Creater      : sadkga sadkga@88.com
 -- @ Since        : 2024-01-22 15:47:29
 -- @ LastAuthor   : sadkga sadkga@88.com
 -- @ LastTime     : 2024-01-22 15:47:29
 -- @ Location     : \\code\\main\\09-FEPAA\\Simba\\DML\\job_stg_fepaa_currency_gc_df.py
 -- @ Message      : 
 -- @ Copyright (c) 2024 by sadkga@88.com, All Rights Reserved. 
 """
# @resource_reference{"/AA/FEPAA/FEPAA_GC_month.csv"}

import pandas as pd
from datetime import datetime
import openpyxl
from azure.storage.blob import ContainerClient
from openpyxl.utils import get_column_letter
import os
import warnings



def get_blob_clib(blobContainName):
    """获取blob客户端,下载数据"""
    connection_string = "DefaultEndpointsProtocol=https;AccountName=proddataplatcn3blob01;AccountKey=ScGueSagWl9s5XDCJeE6xOD8CupGFi5Jp0m9ZVi1Ri812p2GtXD5AXQ/zsVFIcUrNRE2zrIZlWLCjORJZyZHbQ==;EndpointSuffix=core.chinacloudapi.cn"

    container = ContainerClient.from_connection_string(
        conn_str=connection_string,
        container_name=blobContainName)
    return container


def download_local_blob(blob):
    b = container.download_blob(blob=blob)
    blobDirName = os.path.dirname(blob)
    newBlobDirName = os.path.join(blobContainName, blobDirName)
    if not os.path.exists(newBlobDirName):
        os.makedirs(newBlobDirName)
    localFileName = os.path.join(blobContainName, blob)

    with open(localFileName, 'wb') as F:
        f = b.readall()
        F.write(f)
    return localFileName


def data_need(blobs_dir, rn):
    """获取blob目录下的需要的blobs"""
    blobs = list(container.list_blobs(name_starts_with=(blobs_dir)))

    blobs_need = []

    for i in blobs:
        # 获取当前日期的年份和月份
        current_date = datetime.now()
        current_year = current_date.year
        current_month = current_date.month

        # 计算上个月的月份
        previous_month = current_month - 1 if current_month > 1 else 12

        # 确定上个月所在的季度
        previous_quarter = (previous_month - 1) // 3 + 1

        # 如果上个月是一月，则上个季度是去年的第四季度
        if previous_month == 12:
            current_year -= 1
      
        da = i.name
        #  过滤模板
        if da.split('/')[-2] == 'History':
            continue

        if da.split('_')[-2] == rn and int(da.split('.')[0][-1]) == previous_quarter and int(da.split('.')[0][-6:-2]) == current_year:
            blobs_need.append(da)
        else:
            continue

    return blobs_need



def divide_by_100(value):
    """udf,数值/100"""
    return value / 100


def tsrf_LC(path, flile_type):
    """处理GC的xlxs"""
    # 打开Excel文件
    workbook = openpyxl.load_workbook(path, data_only=True)
    # 获取工作表对象
    sheet = workbook.active  # 替换为实际的工作表名称

    # 创建新的工作簿
    new_workbook = openpyxl.Workbook()
    worksheet = new_workbook.active

    # 遍历每一行并写入数据
    for row in sheet.iter_rows():
        first_column_value = row[0].value
        last_three_columns_values = [cell.value for cell in row[-3:]]
        # 将数据写入新的工作表
        worksheet.append([first_column_value] + last_three_columns_values)

    # 检查
    cell_value = worksheet['A3'].value  # 替换为实际的单元格位置
    if flile_type == 'GC':
        # 删除第四行
        worksheet.delete_rows(4)
    elif cell_value == 'PPC-P AA (FEPAA) Status':
         return '1'
  

    df = pd.DataFrame(worksheet.values)
    df = df.iloc[1:]
    df = df.loc[:, df.iloc[1] != 'NO VALUE'].fillna('')
    for i in range(len(df.columns)):
        if df.iloc[0, i] == '':
            # print(LC_history.columns.values[i-1])
            # 如果当前列名为空字符串，则用前一列的列名进行填充
            df.iloc[0, i] = df.iloc[0, i - 1]

    # 重置索引
    df = df.reset_index()

    # 将第一行作为新的列名
    df.columns = df.iloc[0]
    df = df.drop(df.columns[0], axis=1).iloc[1:]
    df_col = list(df.columns)
    df_col.remove('Calendar Year/Month')
    col_set = set(df_col)
    df_list = []


    # 定义缩写月份和对应数字月份的映射关系
    month_mapping = {
        'JAN': '01',
        'FEB': '02',
        'MAR': '03',
        'APR': '04',
        'MAY': '05',
        'JUN': '06',
        'JUL': '07',
        'AUG': '08',
        'SEP': '09',
        'OCT': '10',
        'NOV': '11',
        'DEC': '12'
    }
    for c in col_set:
        df1 = df[['Calendar Year/Month', c]]
        m1 = c.split(' ')
        m = m1[1] + month_mapping[m1[0]]
        df1['calendar_year/month'] = str(m)
        df1['year'] = m1[1]
        df1['month'] = int(month_mapping[m1[0]])
        df1 = df1.iloc[1:]
        df1.columns = ['material_number', 'expierd', 'valid', 'calendar_year/month', 'year',
                       'month']
        for col_name in list(df1.columns):
            df1[col_name] = df1[col_name].astype(str)
        df1 = df1.query('valid != "" or expierd !=""')
        # print(df_filtered)
        df1["is_valid"] = df1.apply(lambda row: 0 if row["expierd"] else 1 if row["valid"] else "", axis=1)
        df_list.append(df1)
        # break

    # 创建一个空的DataFrame作为初始值

    merged_df = pd.concat(df_list)

    merged_df["fepaa"] = merged_df["expierd"].str.strip() + merged_df["valid"].str.strip()
    print(merged_df["fepaa"])

    merged_df["fepaa"] = merged_df["fepaa"].astype(float) / 100
    merged_df["fepaa"] = merged_df["fepaa"].round(4)
    merged_df = pd.DataFrame(merged_df, columns=['material_number', 'calendar_year/month', 'year', 'month', 'fepaa', 'is_valid'])
    merged_df = merged_df.sort_values(by=['material_number', 'calendar_year/month'], ascending=[True, False])
    print(merged_df)
    return merged_df 



if __name__ == '__main__':

    # set
    pd.set_option('display.max_columns', None)  # 显示所有列
    # 忽略SettingWithCopyWarning警告
    warnings.filterwarnings('ignore', category=pd.core.common.SettingWithCopyWarning)


    # todo 1: 定义
    blobContainName = 'bosch-dw-integration-layer'
    blob_dir = 'pig/AA_MBL/FEPAA_Report/'
    save_path = 'AA/FEPAA/FEPAA_GC_month.csv'

    # todo 2: 连接
    container = get_blob_clib(blobContainName)

    # todo 3: 处理
    blobs = data_need(blob_dir, 'GC')
    print(blobs)
    dfs_list = []
    question = []
    rn = 0
    blob_num = len(blobs)
    for blob in blobs:
        rn += 1
        formatted_starttime = datetime.now()

        path = download_local_blob(blob)
        df = tsrf_LC(path, 'GC')
        
          
        formatted_endtime = datetime.now()
        end_time1 = formatted_endtime.strftime("%Y-%m-%d %H:%M:%S")
        time_difference_min = round((formatted_endtime - formatted_starttime).total_seconds()/60, 2)

        print(f'处理文件：{blob}')
        print(f'本次处理时长{time_difference_min}分钟')
        print(f'处理完第{rn}个文件，还有{blob_num - rn}')

        if type(df) == str:
            question.append(blob)
            continue
        elif rn == 1:
            df.to_csv(save_path, mode='w', index=False, header=True)
            
        else:
            df.to_csv(save_path, mode='a', index=False, header=False)

    print(f"格式错误文件{len(question)}个：{question}")
    print('end')
